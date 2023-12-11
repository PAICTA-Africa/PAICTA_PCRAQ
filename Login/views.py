from django.shortcuts import render, redirect
from PIL import Image
import numpy as np
import os
import io
import base64
import uuid
import re
import csv
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.models import User
from django.contrib import messages
from django.template import loader
from django.http import HttpResponse
from .models import Employee
from ipware import get_client_ip
from geopy.geocoders import OpenCage
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template import loader
# from .models import employees
from .models import *
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from django.shortcuts import render
from django.db.models import Sum
from django.views.generic import ListView, DetailView
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import FileResponse
from geopy.geocoders import Nominatim
import time
from datetime import datetime, date
from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from urllib.parse import urlparse, parse_qs
from urllib.parse import urlencode
import socket
from plyer import notification
import io
import openpyxl
from django.http import FileResponse
from datetime import date


def landing(request):
    return render(request, "landing.html", {})
# def detect_face(image_path):
#     """
#     This function takes an image path, detects and returns the face region of the image
#     """
#     image = Image.open(image_path).convert('L')
#     face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
#     faces = face_cascade.detectMultiScale(np.array(image), 1.3, 5)
    
#     if len(faces) == 0:
#         return None
    
#     (x, y, w, h) = faces[0]
#     face = np.array(image)[y:y+h, x:x+w]
#     return face

# def compare_faces(img1, img2):
#     """
#     This function takes two images and returns True if they belong to the same person
#     """
#     face1 = detect_face(img1)
#     face2 = detect_face(img2)
    
#     if face1 is None or face2 is None:
#         return False
    
#     # Resize the images to the same size
#     face1 = cv2.resize(face1, (100, 100))
#     face2 = cv2.resize(face2, (100, 100))
    
#     # Calculate the euclidean distance between the faces
#     dist = np.sqrt(np.sum(np.square(face1 - face2)))
    
#     # If the distance is less than 70, the faces belong to the same person
#     if dist < 70:
#         return True
    
#     return False

def login(request):
    """
    This function handles the login request
    """
    # if request.method == 'POST':
    #     # Get the user's submitted image
    #     submitted_img = request.POST.get('image')
        
    #     # Get the user's saved image from the database
    #     user = User.objects.get(email=request.user.email)
    #     saved_img = user.profile_image
        
    #     # Compare the faces
    #     if compare_faces(submitted_img, saved_img):
    #         # If the faces belong to the same person, unlock
    #         return render(request, 'unlock.html')
    #     else:
    #         # If the faces do not belong to the same person, show error message
    #         messages.error(request, 'Authentication failed.')
            
    return render(request, 'login.html')

def register(request):
    if request.method == 'POST':
        def get_device_ip():
            try:
                # Create a socket connection to a remote host
                sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                sock.connect(("8.8.8.8", 80))  # Connect to a remote host, such as Google DNS

                # Get the local IP address of the socket
                device_ip = sock.getsockname()[0]
                sock.close()

                return device_ip
            except socket.error:
                return None

        device_ip = get_device_ip()
        ip, is_routable = get_client_ip(request)
        if not device_ip:
            device_ip = ip
        email = request.POST['email']
        password = request.POST['password']
        password2 = request.POST['password2']
        name = request.POST['name']
        lastname = request.POST['lastname']
        id_number = request.POST['id_number']
        cell = request.POST['cell']
        dept = request.POST['dept']
        HAddress = request.POST['HAddress']
        WAddress = request.POST['WAddress']
        PAddress = request.POST['PAddress']
        mac_address = hex(uuid.getnode())
        mac_address = mac_address.replace('0x', '')
        mac_address = re.findall('..', mac_address)
        
        if password != password2:
            messages.error(request, "Failed to register: Password characters do not match")
            datatemplate = loader.get_template('authenticate/register.html')
            return HttpResponse(datatemplate.render(context, request))
            exit 
            
        if id_number == "":
            messages.error(request, "Failed to register: ID Number cannot be null")
            datatemplate = loader.get_template('authenticate/register.html')
            return HttpResponse(datatemplate.render(context, request))
            exit 
            
        fpassword = make_password(password)
        post = Employee.objects.create_user(email=email, password=fpassword)
        post.first_name = name
        post.email = email
        post.last_name = lastname
        post.password = fpassword
        post.id_number = id_number
        post.cell = cell
        post.email = email
        post.mobile = cell
        post.dept = dept
        post.home_address = HAddress
        post.work_address = WAddress
        post.postal_address = PAddress
        # post.password = make_password(password, salt="ioutrfucvbuio;)(*&^%)")
        post.ip_address = device_ip
        post.mac_address = mac_address
        post.save()
        
        dtemplate = loader.get_template('authentication/login.html') 
        messages.success(request, "Account created! Login to continue accessing the system.")
        mydata = Employee.objects.all().values()
        context = {
            'mydata': mydata,
            'name': name
        }
        return HttpResponse(dtemplate.render(context, request))
    return render(request, 'authentication/register.html', {})

def Employees(request):
    if request.method == "POST":
        name = request.POST['name']
        department = request.POST['department']
        email = request.POST['email']
        status = request.POST['status']
        project_name = request.POST['project_name']
        role = request.POST['role']
        post = employees()
        post.name = name
        post.department = department
        post.email = email
        post.status = status 
        post.project_name = project_name
        post.project_role = role 
        post.save()
        employ = employees.objects.all()
        return render(request, 'employees.html', {'employ': employ})
    return render(request, 'employees.html')
    
def verify_return(request):
    return render(request, 'authentication/verify.html', {})

def login_user(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        
        try:
            user = Employee.objects.get(email=email)
        except Employee.DoesNotExist:
            messages.error(request, "Login failed: Incorrect user details")
            return render(request, 'authentication/login.html', {})
        
        password_match = check_password(password, user.password)
        
        if password_match:
            if user.is_admin == False:
                if user.is_verified == False:
                    return render(request, 'authentication/verify.html', {'email': user.email})
                else:
                    return render(request, 'index.html', {'user': user})
            elif user.is_admin == True:
                return render(request, 'welcome.html', {'user': user})
            else:
                messages.error(request, "Login failed: Incorrect user type")
                return render(request, 'authentication/login.html', {})
        else:
            messages.error(request, "Login failed: Incorrect user details")
            return render(request, 'authentication/login.html', {})
    
    return render(request, 'authentication/login.html', {})


@login_required
def ClockingView(request):
    employee = request.employee
    return render(request, 'index.html', {'employee': employee})

# @login_required
def clock_action(request, user_id):
    if request.method == "POST":
        def get_notification(title, message):
            notification.notify(
                title=title,
                message=message,
            )
        geolocator = OpenCage(api_key="977d3b34c6db4a13b87930a013e6287e")
        # Location
        def get_location_by_ip():
            try:
                ip_location = geolocator.geocode("East London, South Africa")
                if ip_location is not None:
                    return ip_location.latitude, ip_location.longitude
            except:
                return None, None
        
        #IP Address

        def get_device_ip():
            try:
                # Create a socket connection to a remote host
                sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                sock.connect(("8.8.8.8", 80))  # Connect to a remote host, such as Google DNS

                # Get the local IP address of the socket
                device_ip = sock.getsockname()[0]
                sock.close()

                return device_ip
            except socket.error:
                return None
            
        device_ip = get_device_ip()
        ip, is_routable = get_client_ip(request)
        if device_ip:
            messages.info(request, f"Device IP Address: {device_ip}")
        else:
            device_ip = ip
            messages.error(request, f"Unable to route device")

        IDNumber = request.POST["myid"]
        FirstName = request.POST["name"]
        LastName = request.POST["lastname"]
        latitude, longitude = get_location_by_ip()
        if latitude and longitude:
            final = "{:.6f} {:.6f}".format(latitude, longitude)
            final_loc = geolocator.reverse((latitude, longitude))
        passip = request.POST['passip']
        post = clockin()
        post.user_id = user_id
        post.IDNumber = IDNumber
        post.FirstName = FirstName
        post.LastName = LastName
        post.ip_address = device_ip
        post.geolocation = final_loc
        post.save()
        # if is_routable == 'FALSE':
        #     messages.warning(request, "Successfully clocked in. However, your device is not routable!")
        #     return render(request, "timesheet_success.html", {})
        # else:
        
        try:
            get_notification("PCRAQ - Admin", f"You are clocked in.\nIP: {device_ip}")
        except Exception as e:
            messages.error(request, f"Failed to send notifications. Contact the administration.\n {e}")
        messages.success(request, "Successfully clocked in, your device is routable!")
        mydata = clockin.objects.filter(user_id=user_id)
        context = {
            'mydata': mydata,
            'user_id': user_id,
        }
        
        return render(request, "timesheet_success.html", context)

    return render(request, "index.html", {})

def once_only(request):
    value = 1
    return [value] if value else []   

def take_break(request, dataid, userid):
    now = datetime.now()
    
    # Update the 'takebreak' field of the specific clockin record
    clockin_record = get_object_or_404(clockin, id=dataid, user_id=userid)
    clockin_record.takebreak = now
    clockin_record.save()
    
    messages.success(request, f"Break taken at {now}")
    
    # Load the template and pass the updated data to it
    template = loader.get_template('timesheet_success.html')
    mydata = clockin.objects.filter(user_id=userid)
    context = {
        'mydata': mydata,
        'user_id': userid
    }
    return HttpResponse(template.render(context, request))

def end_break(request, dataid, userid):
    now = datetime.now()
    
    clockin_record = get_object_or_404(clockin, id=dataid, user_id=userid)
    clockin_record.endbreak = now
    clockin_record.save()
    
    messages.success(request, f"Break ended at {now}")
    template = loader.get_template('timesheet_success.html')
    mydata = clockin.objects.filter(user_id=userid)
    context = {
        'mydata': mydata,
        'user_id': userid
    }
    return HttpResponse(template.render(context, request))

def clock_out(request, dataid, userid): 
    now = datetime.now()
    
    clockin_record = get_object_or_404(clockin, id=dataid, user_id=userid)
    clockin_record.clockout_time = now
    clockin_record.save()
    
    messages.success(request, f"Clocked out at {now} successfully!")
    template = loader.get_template('landing.html')
    mydata = Employee.objects.filter(id=userid)
    context = {
        'mydata': mydata,
        'userid': userid
    }
    return HttpResponse(template.render(context, request))
 
def my_profile(request, userid):
    template = loader.get_template('profile.html')
    mydata = Employee.objects.filter(id=userid)
    context = {
        'mydata': mydata,
        'userid': userid
    }

    return HttpResponse(template.render(context, request))

def update_profile(request):
    if request.method == "POST":
        user = request.POST['id']
        fname = request.POST['fname']
        lname = request.POST['lname']
        birth_date = request.POST['birth_date']
        gender = request.POST['gender']
        address = request.POST['address']
        state = request.POST['state']
        country = request.POST['country']
        zip_code = request.POST['zip_code']
        phone_number = request.POST['phone_number']
        department = request.POST['department']
        designation = request.POST['designation']
        report = request.POST['report']
        post = UserProfile()
        post.user_id = user
        post.fname = fname
        post.lname = lname
        post.birth_date = birth_date
        post.gender = gender
        post.address = address
        post.state = state
        post.country = country
        post.zip_code = zip_code
        post.phone_number = phone_number
        post.department = department
        post.designation = designation
        post.report = report
        post.save()
    return render(request, 'profile.html')

def update_info(request):
    if request.method == "POST":
        id_num = request.POST['id_num']
        pass_num = request.POST['pass_num']
        pass_expdate = request.POST['pass_expdate']
        phone_num = request.POST['phone_num']
        nationality = request.POST['nationality']
        religion = request.POST['religion']
        marital_st = request.POST['marital_st']
        emp_spouse = request.POST['emp_spouse']
        num_child = request.POST['num_child']
        post = PersInfo()
        post.id_num = id_num
        post.pass_num = pass_num
        post.pass_expdate = pass_expdate
        post.phone_num = phone_num
        post.nationality = nationality
        post.religion = religion
        post.marital_st = marital_st
        post.emp_spouse = emp_spouse
        post.num_child = num_child
        post.save()
    return render(request, 'profile.html')

def update_emerg(request):
    if request.method == "POST":
        emerg_name = request.POST['emerg_name']
        emerg_rel = request.POST['emerg_rel']
        phone_1 = request.POST['phone_1']
        phone_2 = request.POST['phone_2']
        em_name = request.POST['em_name']
        em_rel = request.POST['em_rel']
        em_phone_1 = request.POST['em_phone_1']
        em_phone_2 = request.POST['em_phone_2']
        post = EmergCon()
        post.emerg_name = emerg_name
        post.emerg_rel = emerg_rel
        post.phone_1 = phone_1
        post.phone_2 = phone_2
        post.em_name = em_name
        post.em_rel = em_rel
        post.em_phone_1 = em_phone_1
        post.em_phone_2 = em_phone_2
        post.save()
    return render(request, 'profile.html')

def update_family_info(request):
    if request.method == "POST":
        fam_member_name = request.POST['fam_member_name']
        fam_rel = request.POST['fam_rel']
        date_of_birth = request.POST['date_of_birth']
        fam_member_phone = request.POST['fam_member_phone']
        post = FamilyInfo()
        post.fam_member_name = fam_member_name
        post.fam_rel = fam_rel
        post.date_of_birth = date_of_birth
        post.fam_member_phone = fam_member_phone
        post.save()
    return render(request, 'profile.html')

def update_education_info(request):
    if request.method == "POST":
        inst = request.POST['inst']
        subject = request.POST['subject']
        start_date = request.POST['start_date']
        compl_date = request.POST['compl_date']
        qual_name = request.POST['qual_name']
        grade = request.POST['grade']
        sec_inst = request.POST['sec_inst']
        sec_subject = request.POST['sec_subject']
        sec_start_date = request.POST['sec_start_date']
        sec_compl_date = request.POST['sec_compl_date']
        sec_qual_name = request.POST['sec_qual_name']
        sec_grade = request.POST['sec_grade']
        post = EducationInfo()
        post.inst = inst
        post.subject = subject
        post.start_date = start_date
        post.compl_date = compl_date
        post.qual_name = qual_name
        post.grade = grade
        post.sec_inst = sec_inst
        post.sec_subject = sec_subject
        post.sec_start_date = sec_start_date
        post.sec_compl_date = sec_compl_date
        post.sec_qual_name = sec_qual_name
        post.sec_grade = sec_grade
        post.save()
    return render(request, 'profile.html')


def update_experience_info(request):
    if request.method == "POST":
        company_name = request.POST['company_name']
        location = request.POST['location']
        job_position = request.POST['job_position']
        period_f = request.POST['period_f']
        period_t = request.POST['period_t']
        sec_company_name = request.POST['sec_company_name']
        sec_location = request.POST['sec_location']
        sec_job_position = request.POST['sec_job_position']
        sec_period_f = request.POST['sec_period_f']
        sec_period_t = request.POST['sec_period_t']
        post = ExperienceInfo()
        post.company_name = company_name
        post.location = location
        post.job_position = job_position
        post.period_f = period_f
        post.period_t = period_t
        post.sec_company_name = sec_company_name
        post.sec_location = sec_location
        post.sec_job_position = sec_job_position
        post.sec_period_f = sec_period_f
        post.sec_period_t = sec_period_t
        post.save()
    return render(request, 'profile.html')

@login_required
def account(request):
    # logout(request)
    return redirect('landing')


def DashboardApp(request, userid):
    mydata = Employee.objects.all().values()
    filter = Employee.objects.filter(id=userid)
    template = loader.get_template('IMS_empDash.html')
    unverified = Employee.objects.filter(is_verified = 0).values()
    context = {
        'mydata': mydata,
        'session': filter,
        'unverified': unverified,
        'userid': userid,
    }

    return HttpResponse(template.render(context, request))
    # return render(request, 'IMS_empDash.html')

def employees_list(request):
    employ = employees.objects.all()
    return render(request, 'employees.html', {'employ': employ})

def registered_users(request, userid):
    mydata = Employee.objects.all().values()
    filter = Employee.objects.all().filter(id=userid)
    template = loader.get_template('registered.html')
    context = {
        'mydata': mydata,
        'session': filter,
        'userid': userid
    }

    return HttpResponse(template.render(context, request))
"""
def leaves(request):
    leave = Leave.objects.all()
    return render(request, 'leaves.html', {'leave':leave})

@receiver(pre_save, sender=Leave)
def update_remaining_number(sender, instance, **kwargs):
    instance.calculate_remaining_number()
"""
        

def resignation(request):
    if request.method == "POST":
        x = resignation_model.objects.all().values()
        template = loader.get_template('resignation.html')
        context = {
            'info': x
        }
        myfiles = request.FILES['resignation_letter']
        fs = FileSystemStorage()
        # data = dict()
        messages.success(request, "Success: You have successfully uploaded your resignation letter!")
        filename = fs.save(myfiles.name, myfiles)
        post = resignation_model()
        post.empname = request.POST['name']
        post.filename = myfiles
        post.reason = request.POST['reason']
        # post.noticedate = request.POST['noticedate']
        # post.resigndate = request.POST['resignation']
        post.save()
        uploaded_file_url = fs.url(filename)
        
        return HttpResponse(template.render(context, request))
        # return render(request, 'resignation.html', {'uploaded_file_url':uploaded_file_url})
    else:
        return render(request, 'resignation.html')

def termination(request):
    if request.method == "POST":
        x = Termination.objects.all().values()
        template = loader.get_template('termination.html')
        context = {
            'info': x
        }
        myfiles = request.FILES['termination_letter']
        fs = FileSystemStorage()
        # data = dict()
        messages.success(request, "Success: You have successfully uploaded your Termination letter!")
        filename = fs.save(myfiles.name, myfiles)
        post = Termination()
        post.empname = request.POST['name']
        post.filename = myfiles
        post.reason = request.POST['reason']
        # post.noticedate = request.POST['noticedate']
        # post.resigndate = request.POST['resignation']
        post.save()
        uploaded_file_url = fs.url(filename)
        
        return HttpResponse(template.render(context, request))
        # return render(request, 'resignation.html', {'uploaded_file_url':uploaded_file_url})
    else:
        return render(request, 'resignation.html')
    
    """
    if request.method == "POST":
        myfiles = request.FILES['termination_letter']
        fs = FileSystemStorage()
        filename = fs.save(myfiles.name, myfiles)
        uploaded_file_url = fs.url(filename)
        
        return render(request, 'termination.html', {'uploaded_file_url':uploaded_file_url})
    else:
        return render(request, 'termination.html')
    """
        
def employee(request):
    template = loader.get_template('IMS_emp.html')
    mydata = Employee.objects.all().values()
    context = {
        'mydata': mydata
    }
    # employ = employees.objects.all()
    return HttpResponse(template.render(context, request))
    

def project_view(request):
    projects= Project.objects.all().values()
    return render(request, 'project_view.html',{'projects':projects})



def leaves_view(request):
    persons = Leave.objects.all()
    context = {'persons': persons}

    return render(request, 'leaves.html', context)

class LeaveListView(ListView):
    model = Leave
    template_name = 'leaves.html'

class LeaveDetailView(DetailView):
    model = Leave
    template_name = 'employee_detail.html'

def leave_summary(request, year):
    leaves = Leave.objects.filter(start_date__year=year)
    leave_counts = {}

    for leave in leaves:
        if leave.leave_type not in leave_counts:
            leave_counts[leave.leave_type] = {
                'total': 0,
                'taken': 0,
                'remaining': 0,
            }

    leave_counts[leave.leave_type]['total'] += leave.duration
    leave_counts[leave.leave_type]['taken'] += leave.duration

    for leave_type, counts in leave_counts.items():
        counts['remaining'] = counts['total'] - counts['taken']

    return render(request, 'leaves.html', {
        'year': year,
        'leave_counts': leave_counts,
    })

def emp_dash(request, userid):

    template = loader.get_template('IMS_emp.html')
    mydata = Employee.objects.filter(id=userid)
    context = {
        'mydata': mydata,
        'user_id': userid
    } 

        # return HttpResponse(template.render(context, request))

    return HttpResponse(template.render(context, request))


def add_project(request):
    if request.method == "POST":
        projectname = request.POST['name']
        leader = request.POST['leader']
        startdate = request.POST['startdate']
        enddate = request.POST['enddate']
        status = request.POST['status']
        version = request.POST['version']
        post = Project()
        post.projectname = projectname
        post.leader = leader
        post.startdate = startdate
        post.enddate = enddate
        post.status = status
        post.github = version
        post.name_id = 0
        post.save()

        messages.success(request, f"Project: {projectname} is successfully added.")
        return render(request, 'IMS_empDash.html', {})
    return render(request, 'IMS_empDash.html', {})

def add_leave(request, userid):
    if request.method == "POST":
        fname = request.POST['fname']
        lname = request.POST['lname']
        startdate = request.POST['startdate']
        enddate = request.POST['enddate']
        reason = request.POST['reason']
        description = request.POST['description']

        str_startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
        str_enddate = datetime.strptime(enddate, '%Y-%m-%d').date()

        num_of_days = (str_enddate - str_startdate).days
        remaining_days = 15 - num_of_days

        if num_of_days > 15:
            messages.error(request, "Leave could not be requested. Days should not be more than 15 days")
            return render(request, "IMS_emp.html", {})
        else:
            post = Leave()
            post.first_name = fname
            post.last_name = lname
            post.start_date = startdate
            post.end_date = enddate
            post.reason = reason
            post.duration = num_of_days
            post.days_left = remaining_days
            post.description = description
            post.userid = userid
            post.save()

            messages.success(request, "Leave is submitted. Awaiting admin approval")
            return render(request, "IMS_emp.html", {})
    return render(request, "IMS_emp.html", {})

def Leave_Request(request):
    user= Employee.objects.get(id=3)
    context = {"user":user}

    if request.method == "POST":
        fname = user.first_name
        lname = user.last_name
        startdate = request.POST['start_date']
        enddate = request.POST['end_date']
        reason = request.POST['reason']
        description = "None"

        str_startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
        str_enddate = datetime.strptime(enddate, '%Y-%m-%d').date()

        num_of_days = (str_enddate - str_startdate).days
        remaining_days = 15 - num_of_days

        if num_of_days > 15:
            messages.error(request, "Leave could not be requested. Days should not be more than 15 days")
            return render(request, "leave_form.html", {})
        else:
            post = Leave()
            post.first_name = fname
            post.last_name = lname
            post.start_date = startdate
            post.end_date = enddate
            post.reason = reason
            post.duration = num_of_days
            post.days_left = remaining_days
            post.description = description
            post.userid = 3
            post.save()

            messages.success(request, "Leave is submitted. Awaiting admin approval")
            return render(request, "leave_form.html", context)
    return render(request, "leave_form.html", context)

# def is_user_in_meeting(user): 
#     current_time = datetime.now()
#     return CalendarEvent.objects.filter(user=user, start_time__lte=current_time, end_time__gte=current_time).exists() 

def upload_document(request, user_id):
    template = loader.get_template('IMS_emp.html')
    mydata = Employee.objects.filter(id=user_id)
    context = {
        'mydata': mydata,
        'user_id': user_id
    } 

        # return HttpResponse(template.render(context, request))
    messages.success(request, "Documents submitted")
    messages.info(request, "Folder created")
    return HttpResponse(template.render(context, request))


def download_registered_users(request):
    # Get All registered users
    users = Employee.objects.all()

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type="text/csv", headers={"Content-Disposition": 'attachment; filename="PCRAQ_users.csv'})
    writer = csv.writer(response)

    # Write the header rows.
    writer.writerow(["id", "last_login", "is_superuser","email", "first_name", "last_name", "id_number", "cell", "dept", "emp_type", "postal", "home_address", "work_address", "postal_address", "ip_address", "mac_address", "reg_time", "is_active", "is_staff", "is_admin", "password"])

    # Write the data rows
    for user in users:
        writer.writerow([user.id,user.last_login,user.is_superuser, user.user.email, user.first_name, user.last_name, user.id_number, user.cell, user.dept, user.emp_type, user.postal, user.home_address, user.work_address, user.postal_address, user.ip_address, user.mac_address, user.reg_time, user.is_active, user.is_staff, user.is_admin, user.password])

    return response

def create_employee(request):
    if request.method == 'POST':
        # Get form data
        email_address = request.POST['email']

        # Create employee instance
        # employee = Employee.objects.create(email=email_address)

        # generate verification token
        token = get_random_string(length=32)

        # Create email verification token
        post = EmailVerificationToken()
        post.email = email_address
        post.token = token
        post.save()

        # Send verification Link to the employee's email
        verification_link = f'https://www.pcraq.com/verify?{urlencode({"token": token, "email": email_address})}'
        subject = "Please verify your email"
        email_body = f"Hi,\n\nPlease click on the following link to verify your email address: {verification_link}\n\nKind Regards\nTeam PCRAQ"
        from_email = "pcraqadmin@paictasa.africa"
        recipient_list = [email_address]

        send_mail(subject, email_body, from_email, recipient_list)

        messages.error(request, "Verification pending!!!")
        # messages.info(request, "Folder created")
        return render(request, 'authentication/login.html')
    return render(request, "auth/verify.html", {})

def verify_email(request):
    token = request.GET.get('token')
    email = request.GET.get('email')
    try:
        email_verification_token = EmailVerificationToken.objects.get(token=token)
    except:
        messages.error("Verification failed")
        return render(request, 'authentication/update.html')

    employee = Employee.objects.get(email=email)
    employee.is_verified = True
    employee.save()

    email_verification_token.delete()
    messages.success(request, "Instance Created: Your email is verified")
    return render(request, "authentication/update.html", {'email':email})


def update_pass(request, email):
    if request.method == 'POST':
        
        password = request.POST['password']
        password2 = request.POST['password2']
        # email = request.GET.get('email')

        if password == password2:
            employee = Employee.objects.filter(email=email).first()
            fpassword = make_password(password)

            employee.password = fpassword
            employee.save()

            messages.success(request, "Password updated. You can now login")

            subject = "Password updated"
            email_body = f"Hi,\n\nPlease note that your password has been updated. You can now access the portal using your new password. \n\nThank you for your understanding.\n\nKind Regards\nTeam PCRAQ"
            from_email = "pcraqadmin@paictasa.africa"
            recipient_list = [email]

            send_mail(subject, email_body, from_email, recipient_list)
            return render(request, "authentication/login.html")
        else:
            messages.error(request, "Failed to update password: Password characters do not match")
            return render(request, "authentication/update.html", {'email': email})


def send(request, userid):
    users = Employee.objects.filter(is_verified=0)
    count = Employee.objects.filter(is_verified=0).count()

    for user in users:
        token = get_random_string(length=32)

        # Create email verification token
        post = EmailVerificationToken()
        post.email = user.email
        post.token = token
        post.save()

        verification_link = f'https://www.pcraq.com/verify?{urlencode({"token": token, "email": user.email})}'
        subject = "Please verify your email"
        email_body = f"Hi {user.first_name},\n\nPlease click on the following link to verify your email address: {verification_link}\n\nKind Regards\nTeam PCRAQ"
        from_email = "pcraqadmin@paictasa.com"
        recipient_list = [user.email]

        send_mail(subject, email_body, from_email, recipient_list)

    mydata = Employee.objects.all().values()
    filter = Employee.objects.filter(id=userid)
    template = loader.get_template('IMS_empDash.html')
    unverified = Employee.objects.filter(is_verified=False)
    context = {
        'mydata': mydata,
        'session': filter,
        'unverified': unverified,
        'userid': userid,
    }

    messages.success(request, f"Token verification has been sent to {count} users")
    return HttpResponse(template.render(context, request))
        # return 
        
def all_clocked(request, userid):
    today_str = date.today().isoformat()
    clocked_users = clockin.objects.filter(clockin_time__startswith=today_str)
    number_of_clocked = clocked_users.count()
    template = loader.get_template('all_clocked.html')
    context = {
        'content': clocked_users,
        'number': number_of_clocked,
        'userid': userid,
    }
    messages.info(request, f"There are {number_of_clocked} clocked users")
    return HttpResponse(template.render(context, request))
    
def download_clocked_users(request):
    # Get All registered users
    today_str = date.today().isoformat()
    users = clockin.objects.filter(clockin_time__startswith=today_str)
    count = users.count()

    # Create a BytesIO buffer to store the Excel content
    buffer = io.BytesIO()

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the filename for the downloaded Excel file
    filename = f"{today_str}_Clocked_Users.xlsx"

    # Set the title of the sheet
    sheet.title = "Clocked Users Report"

    # Add headers to the sheet
    headers = ["First Name", "Last Name", "IP Address", "Location", "Clockin Time", "Break Time", "End Break", "Clockout Time"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    # Add data rows to the sheet
    for row_num, user in enumerate(users, start=2):
        row_values = [
            user.FirstName,
            user.LastName,
            user.ip_address,
            user.geolocation,
            user.clockin_time.replace(tzinfo=None) if user.clockin_time else None,
            user.takebreak.replace(tzinfo=None) if user.takebreak else None,
            user.endbreak.replace(tzinfo=None) if user.endbreak else None,
            user.clockout_time.replace(tzinfo=None) if user.clockout_time else None
        ]
        for col_num, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value

    # Save the Excel content to the buffer
    workbook.save(buffer)

    # Reset the buffer position to read the Excel content
    buffer.seek(0)

    # Create the HTTP response with Excel content and appropriate headers
    response = FileResponse(buffer, as_attachment=True, filename=filename)
    response["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    messages.info(request, f"Report downloaded with {count} users")
    return response

